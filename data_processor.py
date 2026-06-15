import pandas as pd
import openpyxl
import os
import re

# Standard list of 10 main gears (case-insensitive for checks)
STANDARD_GEARS = [
    "bagan perahu",
    "sero",
    "jaring insang tetap",
    "pancing ulur",
    "pukat ikan/pukat hela pertengahan berpapan",
    "rawai dasar",
    "pukat cincin pelagis besar dengan satu kapal",
    "penggaruk tanpa kapal",
    "jala tebar",
    "bubu"
]

def clean_val(val):
    """Clean NaN and null values for clean Excel writing."""
    if pd.isnull(val):
        return ""
    # Convert float representation of integers to integers (like KTP or Handphone numbers)
    if isinstance(val, float) and val.is_integer():
        return int(val)
    return val

def clean_val_str(val):
    """Clean value and force it to string, preserving full numeric IDs (e.g. No. KTP, No. KUSUKA)."""
    if pd.isnull(val):
        return ""
    s = str(val).strip()
    # Remove trailing .0 if the string came from a float representation
    if s.endswith('.0'):
        s = s[:-2]
    return s

def clean_text(text):
    """Normalize text by trimming spaces."""
    if not isinstance(text, str):
        return text
    return text.strip()

def parse_catch_triplet(sp_val, w_val, p_val):
    """
    Parses complex species, weight, and price values.
    Splits multiple values separated by commas, newlines, or 'dan'.
    Returns a list of dictionaries with cleaned keys.
    """
    if pd.isnull(sp_val) or str(sp_val).strip() == "":
        return []
    sp_str = str(sp_val).strip()
    if sp_str.lower() in ["tidak ada", "tidak ada.", "nan", "0", ""]:
        return []
        
    # Split species by newline, comma, or "dan" (case insensitive)
    sp_parts = re.split(r'[\n,\r|]|\bDAN\b|\bdan\b', sp_str)
    species_list = [s.strip().upper() for s in sp_parts if s.strip() and s.strip().lower() not in ["tidak ada", "nan"]]
    
    # Extract numbers from weight and price
    w_str = str(w_val) if pd.notnull(w_val) else ""
    p_str = str(p_val) if pd.notnull(p_val) else ""
    
    # Find all numeric patterns
    weights_raw = re.findall(r'\d+(?:\.\d+)?', w_str)
    prices_raw = re.findall(r'\d+(?:\.\d+)?', p_str)
    
    weights = []
    for x in weights_raw:
        try:
            f = float(x)
            weights.append(int(f) if f.is_integer() else f)
        except ValueError:
            weights.append(x)
            
    prices = []
    for x in prices_raw:
        try:
            f = float(x)
            prices.append(int(f) if f.is_integer() else f)
        except ValueError:
            prices.append(x)
    
    records = []
    for idx, sp in enumerate(species_list):
        # Align weights and prices
        w = weights[idx] if idx < len(weights) else (weights[0] if len(weights) == 1 else "")
        p = prices[idx] if idx < len(prices) else (prices[0] if len(prices) == 1 else "")
        records.append({
            'species': sp,
            'weight': w,
            'price': p
        })
    return records

def process_fisheries_data(input_file_path, template_file_path, output_file_path):
    """
    Reads the Google Form responses Excel file, cleans and maps the data,
    and writes it to the output file using the layout of the template file.
    """
    # Load input data
    df = pd.read_excel(input_file_path, sheet_name=0, dtype={'No. KTP': str, 'No. KUSUKA': str, 'No. HANDPHONE': str})
    
    # Filter out completely empty rows (where NAMA NELAYAN is null)
    df = df[df['NAMA NELAYAN'].notnull()]
    
    # List of Desa columns
    desa_cols = ['DESA', 'DESA 2', 'DESA 3', 'DESA 4', 'DESA 5', 'DESA 6', 'DESA 7', 'DESA 8', 'DESA 9', 'DESA 10', 'DESA 11']
    
    processed_rows = []
    
    for idx, row in df.iterrows():
        record = {}
        
        # 1. Simple direct mappings
        record['Timestamp'] = clean_val(row.get('Timestamp'))
        record['Masukkan Kode Akses'] = clean_val(row.get('Masukkan Kode Akses'))
        record['NAMA PETUGAS '] = clean_val(row.get('Tuliskan Nama Lengkap Anda'))
        record['KECAMATAN'] = clean_val(row.get('KECAMATAN'))
        record['NAMA NELAYAN'] = clean_val(row.get('NAMA NELAYAN'))
        record['No. KTP'] = clean_val_str(row.get('No. KTP'))
        record['No. KUSUKA'] = clean_val_str(row.get('No. KUSUKA'))
        record['No. HANDPHONE'] = clean_val_str(row.get('No. HANDPHONE'))
        record['JENIS USAHA'] = clean_val(row.get('JENIS USAHA'))
        record['KEAHLIAN NELAYAN'] = clean_val(row.get('KEAHLIAN NELAYAN'))
        record['SURAT IZIN BERUSAHA'] = clean_val(row.get('SURAT IZIN BERUSAHA'))
        
        # 2. Merge Desa columns (find first non-null)
        merged_desa = ""
        for d_col in desa_cols:
            if d_col in row and pd.notnull(row[d_col]) and str(row[d_col]).strip() != "":
                merged_desa = str(row[d_col]).strip()
                break
        record['DESA'] = merged_desa
        
        # 3. Ship attributes
        record['NAMA KAPAL'] = clean_val(row.get('NAMA KAPAL'))
        record['STATUS PEMILIKAN KAPAL'] = clean_val(row.get('STATUS PEMILIKAN KAPAL'))
        record['No. SIPI/SIKPI/BPKB'] = clean_val(row.get('No. SIPI/SIKPI/BPKB'))
        record['JUMLAH AWAK KAPAL (ORANG)'] = clean_val(row.get('JUMLAH AWAK KAPAL (ORANG)'))
        
        # Column names for ship details are long in Bank Data
        jenis_kapal_col = [c for c in df.columns if 'JENIS KAPAL' in c and 'Catatan' in c]
        record['JENIS KAPAL'] = clean_val(row.get(jenis_kapal_col[0])) if jenis_kapal_col else ""
        
        record['JUMLAH MESIN'] = clean_val(row.get('JUMLAH MESIN'))
        
        daya_mesin_col = [c for c in df.columns if 'UKURAN DAYA MESIN' in c]
        record['UKURAN DAYA MESIN (PK)'] = clean_val(row.get(daya_mesin_col[0])) if daya_mesin_col else ""
        
        panjang_col = [c for c in df.columns if 'PANJANG KAPAL' in c]
        record['UKURAN PANJANG KAPAL (METER)'] = clean_val(row.get(panjang_col[0])) if panjang_col else ""
        
        lebar_col = [c for c in df.columns if 'LEBAR KAPAL' in c]
        record['UKURAN LEBAR KAPAL (METER)'] = clean_val(row.get(lebar_col[0])) if lebar_col else ""
        
        record['UKURAN GT KAPAL'] = clean_val(row.get('UKURAN GT KAPAL'))
        
        # 4. Gears Mapping (Map as-is from file A to file B without substituting "Yang Lain")
        record['ALAT TANGKAP UTAMA'] = clean_val(row.get('ALAT TANGKAP UTAMA'))
        
        raw_add_gear = str(row.get('ALAT TANGKAP TAMBAHAN', '')).strip()
        if raw_add_gear.lower() in ["tidak ada", "nan", ""]:
            raw_add_gear = ""
        record['ALAT TANGKAP TAMBAHAN'] = raw_add_gear
                
        # 5. Extract Main Catch (Scan triplets 1-11)
        main_jenis, main_berat, main_harga = "", "", ""
        for i in range(1, 12):
            c_col = f"JENIS TANGKAPAN {i}" if i > 1 else "JENIS TANGKAPAN"
            w_col = f"BERAT PER-JENIS TANGKAPAN DALAM 1 KALI TRIP (KG) {i}" if i > 1 else "BERAT PER-JENIS TANGKAPAN DALAM 1 KALI TRIP (KG)"
            p_col = f"HARGA PER-JENIS TANGKAPAN /kg (Rp.) {i}" if i > 1 else "HARGA PER-JENIS TANGKAPAN /kg (Rp.)"
            
            if c_col in row and pd.notnull(row[c_col]) and str(row[c_col]).strip() != "":
                c_val = str(row[c_col]).strip()
                # Ignore if it's explicitly "tidak ada"
                if c_val.lower() not in ["tidak ada", "tidak ada.", "nan", "0"]:
                    main_jenis = c_val
                    main_berat = row.get(w_col)
                    main_harga = row.get(p_col)
                    break
        
        # If no catch was found (or all were empty/ignored), find first non-null anyway
        if not main_jenis:
            for i in range(1, 12):
                c_col = f"JENIS TANGKAPAN {i}" if i > 1 else "JENIS TANGKAPAN"
                w_col = f"BERAT PER-JENIS TANGKAPAN DALAM 1 KALI TRIP (KG) {i}" if i > 1 else "BERAT PER-JENIS TANGKAPAN DALAM 1 KALI TRIP (KG)"
                p_col = f"HARGA PER-JENIS TANGKAPAN /kg (Rp.) {i}" if i > 1 else "HARGA PER-JENIS TANGKAPAN /kg (Rp.)"
                if c_col in row and pd.notnull(row[c_col]):
                    main_jenis = row[c_col]
                    main_berat = row.get(w_col)
                    main_harga = row.get(p_col)
                    break
                    
        # 6. Extract Additional Catch (Triplet 12)
        add_jenis = row.get('JENIS TANGKAPAN 12')
        add_berat = row.get('BERAT PER-JENIS TANGKAPAN DALAM 1 KALI TRIP (KG) 12')
        add_harga = row.get('HARGA PER-JENIS TANGKAPAN /kg (Rp.) 12')
        
        # Parse into separate items
        main_catches = parse_catch_triplet(main_jenis, main_berat, main_harga)
        add_catches = parse_catch_triplet(add_jenis, add_berat, add_harga)
        
        # 7. Other simple columns (will be added to each split row)
        record['JUMLAH TRIP DALAM SEBULAN (HARI)'] = clean_val(row.get('JUMLAH TRIP DALAM SEBULAN (HARI)'))
        record['DAERAH PENANGKAPAN'] = clean_val(row.get('DAERAH PENANGKAPAN'))
        record['KEBUTUHAN OPERASIONAL'] = clean_val(row.get('KEBUTUHAN OPERASIONAL'))
        record['TOTAL BIAYA KESELURUHAN KEBUTUHAN OPERASIONAL (Rp.)'] = clean_val(row.get('TOTAL BIAYA KESELURUHAN KEBUTUHAN OPERASIONAL (Rp.)'))
        record['TEMPAT PENJUALAN HASIL TANGKAPAN '] = clean_val(row.get('TEMPAT PENJUALAN HASIL TANGKAPAN '))
        record['HAMBATAN DAN KENDALA'] = clean_val(row.get('HAMBATAN DAN KENDALA'))
        
        # Split into separate rows if multiple catches exist
        n_rows = max(len(main_catches), len(add_catches))
        if n_rows == 0:
            n_rows = 1  # ensure at least one row is kept
            
        for r_idx in range(n_rows):
            split_record = record.copy()
            
            # Populate main catch for this index
            if r_idx < len(main_catches):
                split_record['JENIS TANGKAPAN'] = main_catches[r_idx]['species']
                split_record['BERAT PER-JENIS TANGKAPAN DALAM 1 KALI TRIP (KG)'] = main_catches[r_idx]['weight']
                split_record['HARGA PER-JENIS TANGKAPAN /kg (Rp.)'] = main_catches[r_idx]['price']
            else:
                split_record['JENIS TANGKAPAN'] = ""
                split_record['BERAT PER-JENIS TANGKAPAN DALAM 1 KALI TRIP (KG)'] = ""
                split_record['HARGA PER-JENIS TANGKAPAN /kg (Rp.)'] = ""
                
            # Populate additional catch for this index
            if r_idx < len(add_catches):
                split_record['JENIS TANGKAPAN 2'] = add_catches[r_idx]['species']
                split_record['BERAT PER-JENIS TANGKAPAN DALAM 1 KALI TRIP (KG) 2'] = add_catches[r_idx]['weight']
                split_record['HARGA PER-JENIS TANGKAPAN /kg (Rp.) 2'] = add_catches[r_idx]['price']
            else:
                split_record['JENIS TANGKAPAN 2'] = ""
                split_record['BERAT PER-JENIS TANGKAPAN DALAM 1 KALI TRIP (KG) 2'] = ""
                split_record['HARGA PER-JENIS TANGKAPAN /kg (Rp.) 2'] = ""
                
            processed_rows.append(split_record)
        
    # Write to template
    wb = openpyxl.load_workbook(template_file_path)
    ws = wb.active
    
    # Identify headers in template to write to correct columns
    headers = [ws.cell(row=1, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
    
    # Clear rows below headers in template
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)
        
    # Write processed data row by row
    for r_idx, record in enumerate(processed_rows, start=2):
        for c_idx, header in enumerate(headers, start=1):
            val = record.get(header, "")
            # Set value
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = val
            
    # Save the updated workbook
    wb.save(output_file_path)
    print(f"Successfully processed {len(processed_rows)} rows and saved to {output_file_path}")
    return pd.DataFrame(processed_rows)

if __name__ == "__main__":
    # Test execution
    process_fisheries_data(
        "BANK DATA PERIKANAN TANGKAP (Jawaban).xlsx",
        "template data input.xlsx",
        "analisa_download.xlsx"
    )
